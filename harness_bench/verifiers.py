"""Small helpers for writing task verifiers.

All helpers return a `Verifier` (`Callable[[Path], VerifyResult]`) so they
can be composed via `all_of(...)`.
"""

from __future__ import annotations

import json
import re
import sqlite3
import subprocess
import sys
from pathlib import Path
from typing import Any

from harness_bench.core import Verifier, VerifyResult


def _read_utf8_text(path: Path, rel: str) -> str | VerifyResult:
    """Read a benchmark text artifact as UTF-8 or return a clean failure.

    Benchmark outputs are expected to be UTF-8. Be explicit so Windows locale
    defaults (cp1251/cp866/charmap) cannot change verifier behaviour, but do not
    let a UnicodeDecodeError escape as a verifier traceback.
    """
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        return VerifyResult(False, f"{rel} is not valid UTF-8: {exc}")


def file_exists(rel: str) -> Verifier:
    """Pass when `rel` exists as a regular file."""

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if p.exists() and p.is_file():
            return VerifyResult(True, f"{rel} exists")
        return VerifyResult(False, f"{rel} missing")

    return _check


def file_does_not_exist(rel: str) -> Verifier:
    """Pass when `rel` is absent."""

    def _check(ws: Path) -> VerifyResult:
        if not (ws / rel).exists():
            return VerifyResult(True, f"{rel} absent")
        return VerifyResult(False, f"{rel} still exists")

    return _check


def file_text_equals(rel: str, expected: str, *, strip: bool = True) -> Verifier:
    """Pass when the contents of `rel` equal `expected` (optionally stripped)."""

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        actual = _read_utf8_text(p, rel)
        if isinstance(actual, VerifyResult):
            return actual
        a, e = (actual.strip(), expected.strip()) if strip else (actual, expected)
        if a == e:
            return VerifyResult(True, f"{rel} matches expected content")
        return VerifyResult(False, f"{rel} content differs\nExpected: {e!r}\nActual:   {a!r}")

    return _check


def file_contains(rel: str, *needles: str) -> Verifier:
    """Pass when every `needle` is found in the file's text."""

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        text = _read_utf8_text(p, rel)
        if isinstance(text, VerifyResult):
            return text
        missing = [n for n in needles if n not in text]
        if missing:
            return VerifyResult(False, f"{rel} missing substrings: {missing!r}")
        return VerifyResult(True, f"{rel} contains all expected substrings")

    return _check


def file_not_contains(rel: str, *needles: str) -> Verifier:
    """Pass when none of the `needles` are present in the file's text."""

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        text = _read_utf8_text(p, rel)
        if isinstance(text, VerifyResult):
            return text
        present = [n for n in needles if n in text]
        if present:
            return VerifyResult(False, f"{rel} still contains forbidden: {present!r}")
        return VerifyResult(True, f"{rel} free of forbidden substrings")

    return _check


def file_matches_regex(rel: str, pattern: str, *, flags: int = re.MULTILINE) -> Verifier:
    """Pass when the file contents match `pattern` (Python regex)."""

    rx = re.compile(pattern, flags)

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        text = _read_utf8_text(p, rel)
        if isinstance(text, VerifyResult):
            return text
        if rx.search(text):
            return VerifyResult(True, f"{rel} matches /{pattern}/")
        return VerifyResult(False, f"{rel} does not match /{pattern}/")

    return _check


def file_lines_equal(rel: str, expected_lines: list[str]) -> Verifier:
    """Pass when the file's non-empty lines equal `expected_lines` exactly."""

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        text = _read_utf8_text(p, rel)
        if isinstance(text, VerifyResult):
            return text
        actual = [line for line in text.splitlines() if line.strip() != ""]
        if actual == expected_lines:
            return VerifyResult(True, f"{rel} lines match")
        return VerifyResult(
            False,
            f"{rel} lines differ\nExpected: {expected_lines!r}\nActual:   {actual!r}",
        )

    return _check


def json_file_has(rel: str, **expected_pairs: Any) -> Verifier:
    """Pass when `rel` parses as JSON and each key/value pair matches."""

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        try:
            text = _read_utf8_text(p, rel)
            if isinstance(text, VerifyResult):
                return text
            data = json.loads(text)
        except json.JSONDecodeError as exc:
            return VerifyResult(False, f"{rel} invalid JSON: {exc}")
        if not isinstance(data, dict):
            return VerifyResult(False, f"{rel} is not a JSON object (got {type(data).__name__})")
        wrong: list[str] = []
        for key, expected in expected_pairs.items():
            actual = data.get(key, _MISSING)
            if actual != expected:
                wrong.append(f"{key}={actual!r} (expected {expected!r})")
        if wrong:
            return VerifyResult(False, f"{rel} mismatch: {'; '.join(wrong)}")
        return VerifyResult(True, f"{rel} has expected JSON keys")

    return _check


def json_file_matches(rel: str, expected: Any) -> Verifier:
    """Pass when `rel` parses as JSON and equals `expected` exactly."""

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        try:
            text = _read_utf8_text(p, rel)
            if isinstance(text, VerifyResult):
                return text
            data = json.loads(text)
        except json.JSONDecodeError as exc:
            return VerifyResult(False, f"{rel} invalid JSON: {exc}")
        if data == expected:
            return VerifyResult(True, f"{rel} matches expected JSON")
        return VerifyResult(False, f"{rel} JSON mismatch\nExpected: {expected!r}\nActual:   {data!r}")

    return _check


def python_runs(
    rel: str,
    *,
    expected_stdout: str | None = None,
    expected_substring: str | None = None,
    timeout: int = 20,
) -> Verifier:
    """Run `python rel` as a subprocess in the workspace and check output."""

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        try:
            result = subprocess.run(  # noqa: S603 — trusted local benchmark
                [sys.executable, str(p)],
                cwd=ws,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=timeout,
                check=False,
            )
        except subprocess.TimeoutExpired:
            return VerifyResult(False, f"{rel} timed out after {timeout}s")
        if result.returncode != 0:
            return VerifyResult(
                False,
                f"{rel} exit_code={result.returncode}; stderr: {result.stderr.strip()[:400]}",
            )
        out = result.stdout
        if expected_stdout is not None and out.rstrip("\n") != expected_stdout.rstrip("\n"):
            return VerifyResult(
                False,
                f"{rel} stdout mismatch\nExpected: {expected_stdout!r}\nActual:   {out!r}",
            )
        if expected_substring is not None and expected_substring not in out:
            return VerifyResult(False, f"{rel} stdout missing {expected_substring!r}; got {out!r}")
        return VerifyResult(True, f"{rel} ran and produced expected output")

    return _check


def python_callable_returns(rel: str, call_expr: str, expected: Any) -> Verifier:
    """Import the module at `rel` as `mod` and assert `<call_expr>` equals `expected`.

    Example: `python_callable_returns("greeting.py", "mod.greet('Bob')", "Hi, Bob!")`.
    """

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        snippet = (
            "import importlib.util as u, json, sys\n"
            f"spec = u.spec_from_file_location('mod', r'{p}')\n"
            "mod = u.module_from_spec(spec); sys.modules['mod'] = mod\n"
            "spec.loader.exec_module(mod)\n"
            f"val = {call_expr}\n"
            "sys.stdout.write(json.dumps(val, ensure_ascii=False, default=str))\n"
        )
        try:
            result = subprocess.run(  # noqa: S603 — trusted local benchmark
                [sys.executable, "-c", snippet],
                cwd=ws,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=20,
                check=False,
            )
        except subprocess.TimeoutExpired:
            return VerifyResult(False, f"{rel} call timed out")
        if result.returncode != 0:
            return VerifyResult(
                False,
                f"{rel} runtime error: {result.stderr.strip()[:400]}",
            )
        try:
            actual = json.loads(result.stdout)
        except json.JSONDecodeError:
            return VerifyResult(False, f"{rel} produced non-JSON output: {result.stdout!r}")
        if actual == expected:
            return VerifyResult(True, f"{rel}: {call_expr} == {expected!r}")
        return VerifyResult(
            False, f"{rel}: {call_expr} returned {actual!r}, expected {expected!r}"
        )

    return _check


def all_of(*verifiers: Verifier) -> Verifier:
    """Compose verifiers; pass only when all pass."""

    def _check(ws: Path) -> VerifyResult:
        messages: list[str] = []
        for v in verifiers:
            r = v(ws)
            if not r.passed:
                return VerifyResult(False, r.message)
            messages.append(r.message)
        return VerifyResult(True, "; ".join(messages))

    return _check


def pytest_passes(test_dir: str = "tests", *, timeout: int = 60) -> Verifier:
    """Verifier: run `python -m pytest -q <test_dir>` and require exit 0."""

    def _check(ws: Path) -> VerifyResult:
        try:
            result = subprocess.run(  # noqa: S603 — benchmark only
                [sys.executable, "-m", "pytest", "-q", test_dir],
                cwd=ws,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=timeout,
                check=False,
            )
        except subprocess.TimeoutExpired:
            return VerifyResult(False, f"pytest timed out after {timeout}s")
        if result.returncode == 0:
            return VerifyResult(True, f"pytest {test_dir}/ passed")
        tail = (result.stdout + result.stderr).strip()[-400:]
        return VerifyResult(False, f"pytest exit={result.returncode}; tail: {tail!r}")

    return _check


def xlsx_cell_equals(rel: str, sheet: str, cell: str, expected: Any) -> Verifier:
    """Verifier: assert that an xlsx cell equals the expected value.

    Numbers compared loosely (so `999` and `999.0` both pass when the gold is
    an integer). Strings compared exactly.
    """

    def _check(ws: Path) -> VerifyResult:
        import openpyxl  # noqa: PLC0415 — heavy import; only when verifier runs

        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        wb = openpyxl.load_workbook(p, data_only=True)
        if sheet not in wb.sheetnames:
            return VerifyResult(False, f"{rel} has no sheet {sheet!r}; sheets={wb.sheetnames}")
        value = wb[sheet][cell].value
        if value == expected:
            return VerifyResult(True, f"{rel}[{sheet}!{cell}] == {expected!r}")
        try:
            if float(value) == float(expected):
                return VerifyResult(True, f"{rel}[{sheet}!{cell}] ≈ {expected!r}")
        except (TypeError, ValueError):
            pass
        return VerifyResult(False, f"{rel}[{sheet}!{cell}] = {value!r}, expected {expected!r}")

    return _check


def sqlite_query_returns(rel: str, query: str, expected: Any) -> Verifier:
    """Verifier: run a query on a sqlite db and compare the first cell.

    Useful for "count rows", "sum a column", "look up a value by id" tasks.
    """

    def _check(ws: Path) -> VerifyResult:
        p = ws / rel
        if not p.exists():
            return VerifyResult(False, f"{rel} missing")
        try:
            conn = sqlite3.connect(p)
            try:
                row = conn.execute(query).fetchone()
            finally:
                conn.close()
        except sqlite3.Error as exc:
            return VerifyResult(False, f"{rel} sqlite error: {exc}")
        if row is None:
            return VerifyResult(False, f"{rel}: query returned no rows")
        value = row[0]
        if value == expected:
            return VerifyResult(True, f"{rel}: {query!s} == {value!r}")
        return VerifyResult(False, f"{rel}: {query!s} = {value!r}, expected {expected!r}")

    return _check


_MISSING = object()
